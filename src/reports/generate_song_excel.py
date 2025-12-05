#!/usr/bin/env python3
"""
Generate Excel spreadsheet with separate sheets for each shared song
Each sheet contains all TikTok video links where that song was used
"""

import sys
import subprocess
import json
from datetime import datetime
from collections import defaultdict
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

# Import configuration
try:
    import config
except ImportError:
    print("ERROR: config.py not found. Please create config.py with your account configuration.")
    sys.exit(1)

# Use accounts and exclusive songs from config
ACCOUNTS = config.ACCOUNTS
EXCLUSIVE_SONGS = config.EXCLUSIVE_SONGS

def scrape_account(account):
    """Scrape a single TikTok account using yt-dlp - ALL videos"""
    print(f"Scraping {account}...")

    profile_url = f"https://www.tiktok.com/{account}"

    cmd = [
        config.YT_DLP_CMD,
        '--flat-playlist',
        '--dump-json',
        profile_url
    ]

    result = subprocess.run(cmd, capture_output=True, text=True)

    if result.returncode != 0:
        print(f"  ⚠️  Error scraping {account}: {result.stderr.strip()}")
        return []

    videos = []
    for line in result.stdout.strip().split('\n'):
        if line:
            try:
                video_data = json.loads(line)

                # Extract upload date
                upload_date = video_data.get('upload_date', '')
                if upload_date:
                    try:
                        upload_datetime = datetime.strptime(upload_date, '%Y%m%d')

                        # Filter for configured cutoff date
                        if upload_datetime < config.CUTOFF_DATE:
                            continue

                        formatted_date = upload_datetime.strftime('%Y-%m-%d')
                    except:
                        formatted_date = upload_date
                else:
                    formatted_date = 'Unknown'
                    continue  # Skip if no date

                song_title = video_data.get('track', '') or 'Unknown'
                song_artist = video_data.get('artist', '') or video_data.get('creator', '') or 'Unknown'
                sound_key = f"{song_title} - {song_artist}"

                # Skip exclusive songs
                if sound_key in EXCLUSIVE_SONGS:
                    continue

                video_url = video_data.get('webpage_url') or video_data.get('url', '')

                videos.append({
                    'account': account,
                    'sound_key': sound_key,
                    'song_title': song_title,
                    'song_artist': song_artist,
                    'url': video_url,
                    'upload_date': formatted_date,
                    'views': video_data.get('view_count', 0),
                    'likes': video_data.get('like_count', 0),
                    'comments': video_data.get('comment_count', 0),
                    'shares': video_data.get('repost_count', 0)
                })

            except json.JSONDecodeError:
                continue

    cutoff_date_str = config.CUTOFF_DATE.strftime('%Y-%m-%d')
    print(f"  ✓ Found {len(videos)} videos (shared songs, from {cutoff_date_str} onwards)")
    return videos

def create_excel_workbook(videos_by_song):
    """Create Excel workbook with separate sheet for each song"""
    wb = openpyxl.Workbook()
    wb.remove(wb.active)  # Remove default sheet

    # Sort songs alphabetically
    sorted_songs = sorted(videos_by_song.items(), key=lambda x: x[0])

    for song_key, videos in sorted_songs:
        # Create sheet name (max 31 chars for Excel)
        sheet_name = song_key[:31] if len(song_key) <= 31 else song_key[:28] + "..."

        # Make sheet name Excel-safe (no special chars)
        sheet_name = sheet_name.replace('/', '-').replace('\\', '-').replace('?', '').replace('*', '').replace('[', '').replace(']', '').replace(':', '-')

        ws = wb.create_sheet(title=sheet_name)

        # Header row
        headers = ['Account', 'TikTok URL', 'Upload Date', 'Views', 'Likes', 'Comments', 'Shares']
        ws.append(headers)

        # Style header row
        header_fill = PatternFill(start_color='4F81BD', end_color='4F81BD', fill_type='solid')
        header_font = Font(bold=True, color='FFFFFF')

        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_num)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center', vertical='center')

        # Add video data
        for video in videos:
            ws.append([
                video['account'],
                video['url'],
                video['upload_date'],
                video['views'],
                video['likes'],
                video['comments'],
                video['shares']
            ])

        # Adjust column widths
        ws.column_dimensions['A'].width = 30  # Account
        ws.column_dimensions['B'].width = 50  # URL
        ws.column_dimensions['C'].width = 15  # Date
        ws.column_dimensions['D'].width = 12  # Views
        ws.column_dimensions['E'].width = 12  # Likes
        ws.column_dimensions['F'].width = 12  # Comments
        ws.column_dimensions['G'].width = 12  # Shares

        # Freeze header row
        ws.freeze_panes = 'A2'

    return wb

def main():
    print("\n" + "="*80)
    print("GENERATING EXCEL SPREADSHEET - SHARED SONGS BY TikTok USAGE")
    print("="*80)
    cutoff_date_str = config.CUTOFF_DATE.strftime('%Y-%m-%d')
    print(f"\nScraping {len(ACCOUNTS)} accounts for shared song usage (from {cutoff_date_str} onwards)...\n")

    all_videos = []

    for account in ACCOUNTS:
        videos = scrape_account(account)
        all_videos.extend(videos)

    print(f"\n{'='*80}")
    print(f"Total videos collected: {len(all_videos)}")
    print(f"{'='*80}\n")

    if not all_videos:
        print("⚠️  No videos found. Exiting.")
        return

    # Group videos by song
    videos_by_song = defaultdict(list)
    for video in all_videos:
        videos_by_song[video['sound_key']].append(video)

    # Sort videos within each song by date (newest first)
    for song_key in videos_by_song:
        videos_by_song[song_key].sort(key=lambda x: x['upload_date'], reverse=True)

    print(f"Found {len(videos_by_song)} unique shared songs\n")
    print("Creating Excel workbook...")

    wb = create_excel_workbook(videos_by_song)

    output_file = config.EXCEL_OUTPUT_FILE
    wb.save(str(output_file))

    print(f"✓ Excel file generated: {output_file}")
    print(f"  - {len(wb.worksheets)} sheets (one per song)")
    print(f"  - {len(all_videos)} total video links")
    print(f"\n{'='*80}\n")

if __name__ == '__main__':
    main()
